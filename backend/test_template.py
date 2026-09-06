import csv
import io
import openpyxl

headers = [
    "First Name", "Middle Initial", "Last Name", "Date of Birth", "Gender",
    "Phone Number", "Email Address", "Street Address", "Province", "City",
    "Postal Code", "Emergency Contact Name", "Emergency Contact Phone",
    "Emergency Contact Relationship"
]

def generate_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    return output.getvalue()

def generate_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    ws.append(headers)
    
    # Optional styling
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

print("CSV generated:", len(generate_csv()), "bytes")
print("XLSX generated:", len(generate_xlsx()), "bytes")
